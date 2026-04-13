"""End-to-end test for the Modal cleaning API.

Uses the repo's dirty-data generator to create synthetic input, runs the
full pipeline via the FastAPI TestClient with a mock LLM responder, and
verifies:
  • Job reaches 'done' state
  • All 4 output files exist
  • Color coding is applied in cleaned_data.xlsx
  • Convergence trace is recorded in report_summary.json
"""

from __future__ import annotations

import json
import math
import os
import sys
import tempfile
from typing import Any, Dict, List

import pytest
from fastapi.testclient import TestClient

# ---------------------------------------------------------------------------
# Ensure the modal-api directory is on sys.path so the fallback absolute
# imports inside app.py / pipeline.py work.
# ---------------------------------------------------------------------------

MODAL_API_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT = os.path.dirname(MODAL_API_DIR)

if MODAL_API_DIR not in sys.path:
    sys.path.insert(0, MODAL_API_DIR)
if REPO_ROOT not in sys.path:
    sys.path.insert(0, REPO_ROOT)

# Patch the volume mount to a temp directory BEFORE importing app
_tmpdir = tempfile.mkdtemp(prefix="scp_test_")

import job_store  # type: ignore[import-untyped]
import pipeline as pipeline_mod  # type: ignore[import-untyped]
import output_builder  # type: ignore[import-untyped]
import app as app_module  # type: ignore[import-untyped]

app_module.VOLUME_MOUNT = _tmpdir

# Reset the in-process job store
job_store.reset_store()

client = TestClient(app_module.web_app)

# ---------------------------------------------------------------------------
# Synthetic dirty data generator (inline, adapted from generate_dirty.py)
# ---------------------------------------------------------------------------

SAMPLE_TAXONOMY = {
    "Energy & Fuel": {
        "Diesel": ["Bulk diesel delivery", "Diesel storage"],
        "Electricity": ["Grid electricity", "On-site generation"],
    },
    "Equipment & Parts": {
        "Mobile Equipment": ["Haul truck parts", "Excavator parts"],
        "Fixed Plant": ["Conveyor components", "Crusher parts"],
    },
    "Maintenance & Repair": {
        "Planned Maintenance": ["Lubrication services", "Shutdown services"],
        "Breakdown & Emergency": ["Emergency repair", "Welding services"],
    },
    "Professional Services": {
        "IT & Technology": ["Cloud services", "Software licences"],
        "Legal & Advisory": ["Audit services", "Legal counsel"],
    },
}

SAMPLE_VENDOR_DICT = {
    "Cat": "Caterpillar",
    "CAT": "Caterpillar",
    "Caterpiller": "Caterpillar",
    "KOMATSU": "Komatsu",
    "Komatsu Aust": "Komatsu",
    "Sandvik": "Sandvik Mining",
    "SANDVIK": "Sandvik Mining",
    "ABB": "ABB Ltd",
    "Siemens": "Siemens Ltd",
    "SIEMENS": "Siemens Ltd",
}

SAMPLE_ABBREV_DICT = {
    "Cat": "Caterpillar",
    "maint": "maintenance",
    "svc": "service",
    "equip": "equipment",
}


def _generate_dirty_rows(n: int = 60) -> List[Dict[str, Any]]:
    """Generate synthetic dirty procurement rows for testing."""
    import random
    random.seed(42)

    suppliers_dirty = [
        "Cat", "CAT", "Caterpiller", "KOMATSU", "Komatsu Aust",
        "Sandvik", "SANDVIK", "ABB", "SIEMENS", "UnknownVendorXYZ",
        "Totally Fake Co", "Mystery Supplier",
    ]

    date_formats = [
        "04/07/2025", "2025-07-04", "4-Jul-25", "20250704",
        "07/04/2025", "45842", "2025-01-15", "15/01/2025",
        "bad-date", "", "2024-12-31",
    ]

    amount_formats = [
        "$1,500.00", "AUD 2500", "1.500,00", "750", "(1200.50)",
        "3000.00", "$45,678.90", "invalid", "", "100.00",
    ]

    descriptions = [
        "Bulk diesel delivery for site",
        "Haul truck parts replacement",
        "Emergency repair svc",
        "Cloud services subscription",
        "Lubrication maint equip",
        "Conveyor components",
        "Software licences renewal",
        "Audit services Q3",
        "Welding svc on crusher",
        "Grid electricity monthly",
    ]

    units = ["litres", "pcs", "hours", "each", "tonne", "sqm", "lump sum", "kg", "EA", "unknown_unit"]

    rows = []
    for i in range(n):
        has_amount = random.random() > 0.15
        has_qty = random.random() > 0.15
        has_up = random.random() > 0.2

        amount_val = random.choice(amount_formats) if has_amount else None
        qty_val = round(random.uniform(1, 500), 2) if has_qty else None
        up_val = round(random.uniform(0.5, 100), 2) if has_up else None

        row = {
            "record_id": f"TEST-{i+1:05d}",
            "date": random.choice(date_formats),
            "financial_year": "",
            "invoice_number": f"INV-{random.randint(1000, 9999)}",
            "description": random.choice(descriptions) + ("\ufeff" if random.random() < 0.05 else ""),
            "quantity": qty_val,
            "unit": random.choice(units),
            "unit_price": up_val,
            "amount": amount_val,
            "currency": random.choice(["AUD", "USD"]),
            "supplier_name": random.choice(suppliers_dirty),
            "supplier_id": "",
            "category_l1": random.choice(["", "Energy & Fuel", "Unknown Category", "Equipment & Parts"]),
            "category_l2": "",
            "category_l3": "",
        }
        rows.append(row)

    return rows


# ---------------------------------------------------------------------------
# Mock LLM responder
# ---------------------------------------------------------------------------


def mock_llm_respond(batch: List[dict]) -> List[dict]:
    """Deterministic mock LLM that provides cleaned values with confidence."""
    responses = []
    for item in batch:
        row_id = item["row_id"]
        fields = item.get("fields_needing_reasoning", [])
        current = item.get("current_values", {})
        reasons = item.get("reasons", {})

        field_values = {}
        confidence = {}

        for f in fields:
            if f == "supplier_name":
                dirty = current.get("supplier_name", "")
                if "unknown" in dirty.lower() or "fake" in dirty.lower() or "mystery" in dirty.lower():
                    field_values[f] = dirty
                    confidence[f] = 0.5
                else:
                    field_values[f] = dirty
                    confidence[f] = 0.9
            elif f == "category_l1":
                desc = str(current.get("description", "")).lower()
                if "diesel" in desc or "electricity" in desc or "fuel" in desc:
                    field_values[f] = "Energy & Fuel"
                    confidence[f] = 0.9
                elif "truck" in desc or "conveyor" in desc or "crusher" in desc:
                    field_values[f] = "Equipment & Parts"
                    confidence[f] = 0.85
                elif "repair" in desc or "maint" in desc or "welding" in desc or "lubrication" in desc:
                    field_values[f] = "Maintenance & Repair"
                    confidence[f] = 0.8
                elif "cloud" in desc or "software" in desc or "audit" in desc or "legal" in desc:
                    field_values[f] = "Professional Services"
                    confidence[f] = 0.85
                else:
                    field_values[f] = "Equipment & Parts"
                    confidence[f] = 0.6
            elif "amount" in f or "quantity" in f or "unit_price" in f:
                field_values[f] = current.get(f)
                confidence[f] = 0.7
            else:
                field_values[f] = current.get(f)
                confidence[f] = 0.75

        responses.append({
            "row_id": row_id,
            "field_values": field_values,
            "confidence_per_field": confidence,
            "notes": "Mock LLM response",
        })

    return responses


# ===========================================================================
# Tests
# ===========================================================================


class TestEndToEnd:
    """Full end-to-end test exercising the polling API pattern."""

    def setup_method(self):
        job_store.reset_store()

    def test_full_pipeline(self):
        dirty_rows = _generate_dirty_rows(60)

        # 1. POST /jobs
        resp = client.post("/jobs", json={
            "data": dirty_rows,
            "taxonomy": SAMPLE_TAXONOMY,
            "abbreviation_dictionary": SAMPLE_ABBREV_DICT,
            "vendor_dictionary": SAMPLE_VENDOR_DICT,
            "max_iterations": 5,
            "llm_batch_size": 25,
        })
        assert resp.status_code == 200, f"Create job failed: {resp.text}"
        body = resp.json()
        job_id = body["job_id"]
        assert job_id
        assert body["total_rows"] == 60

        # 2. Poll status
        status_resp = client.get(f"/jobs/{job_id}/status")
        assert status_resp.status_code == 200
        status = status_resp.json()
        assert status["state"] in ("awaiting_llm_batch", "done")

        # 3. LLM loop
        max_loops = 10
        loop_count = 0

        while loop_count < max_loops:
            loop_count += 1
            status_resp = client.get(f"/jobs/{job_id}/status")
            state = status_resp.json()["state"]

            if state == "done":
                break

            if state == "error":
                pytest.fail(f"Job errored: {status_resp.json()}")

            if state != "awaiting_llm_batch":
                continue

            # Get next batch
            batch_resp = client.get(f"/jobs/{job_id}/next_llm_batch")
            if batch_resp.status_code == 204:
                continue

            assert batch_resp.status_code == 200
            batch_data = batch_resp.json()
            batch = batch_data["batch"]

            if not batch:
                continue

            # Mock LLM responds
            llm_answers = mock_llm_respond(batch)

            # Submit responses
            submit_resp = client.post(
                f"/jobs/{job_id}/llm_batch_response",
                json={"responses": llm_answers},
            )
            assert submit_resp.status_code == 200
            submit_body = submit_resp.json()

            if submit_body.get("converged"):
                # Wait for finalization
                status_resp = client.get(f"/jobs/{job_id}/status")
                break

        # 4. Verify job is done
        final_status = client.get(f"/jobs/{job_id}/status").json()
        assert final_status["state"] == "done", f"Job did not reach done: {final_status}"

        # 5. Get results
        results_resp = client.get(f"/jobs/{job_id}/results")
        assert results_resp.status_code == 200
        results = results_resp.json()

        # Verify all 4 output files referenced
        signed_urls = results["signed_urls"]
        assert "cleaned_data" in signed_urls
        assert "changes_only" in signed_urls
        assert "user_verification_required" in signed_urls
        assert "report_summary" in signed_urls

        # Verify files actually exist on disk
        job = job_store.get_job(job_id)
        output_paths = job["output_paths"]
        for name, path in output_paths.items():
            assert os.path.exists(path), f"Output file missing: {path}"

        # 6. Verify color coding in cleaned_data.xlsx
        self._verify_color_coding(output_paths["cleaned_data"])

        # 7. Verify convergence trace in report_summary.json
        self._verify_report_summary(output_paths["report_summary"])

        # 8. Verify changes_only has content
        self._verify_changes_only(output_paths["changes_only"])

    def _verify_color_coding(self, path: str):
        from openpyxl import load_workbook
        wb = load_workbook(path)
        ws = wb.active

        fills_found = {"green": False, "yellow": False, "red": False}

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.fill and cell.fill.start_color:
                    color = str(cell.fill.start_color.rgb) if cell.fill.start_color.rgb else ""
                    if "B6D7A8" in color.upper():
                        fills_found["green"] = True
                    elif "FFF2CC" in color.upper():
                        fills_found["yellow"] = True
                    elif "E06666" in color.upper():
                        fills_found["red"] = True

        assert fills_found["green"], "No green (changed) cells found in cleaned_data.xlsx"
        # Yellow and red may or may not appear depending on data

    def _verify_report_summary(self, path: str):
        with open(path) as f:
            summary = json.load(f)

        assert "total_rows" in summary
        assert summary["total_rows"] == 60
        assert "convergence_trace" in summary
        assert len(summary["convergence_trace"]) > 0
        assert "rule_stats" in summary
        assert "pass_timings" in summary

    def _verify_changes_only(self, path: str):
        from openpyxl import load_workbook
        wb = load_workbook(path)
        ws = wb.active
        assert ws.max_row > 1, "changes_only.xlsx should have at least one data row"

    def test_job_not_found(self):
        resp = client.get("/jobs/nonexistent/status")
        assert resp.status_code == 404

    def test_results_before_done(self):
        dirty = _generate_dirty_rows(5)
        resp = client.post("/jobs", json={
            "data": dirty,
            "taxonomy": SAMPLE_TAXONOMY,
            "vendor_dictionary": SAMPLE_VENDOR_DICT,
        })
        job_id = resp.json()["job_id"]
        state = client.get(f"/jobs/{job_id}/status").json()["state"]
        if state != "done":
            results_resp = client.get(f"/jobs/{job_id}/results")
            assert results_resp.status_code == 409

    def test_no_flagged_rows_goes_straight_to_done(self):
        """If all rows can be cleaned deterministically, job goes to done."""
        clean_rows = [
            {
                "record_id": "CLEAN-001",
                "date": "2025-07-04",
                "amount": 1500.0,
                "quantity": 10,
                "unit_price": 150.0,
                "unit": "EA",
                "supplier_name": "Caterpillar",
                "category_l1": "Equipment & Parts",
                "category_l2": "Mobile Equipment",
                "category_l3": "Haul truck parts",
                "description": "Clean data row",
            }
        ]
        resp = client.post("/jobs", json={
            "data": clean_rows,
            "taxonomy": SAMPLE_TAXONOMY,
            "vendor_dictionary": {"Caterpillar": "Caterpillar"},
        })
        assert resp.status_code == 200
        job_id = resp.json()["job_id"]
        status = client.get(f"/jobs/{job_id}/status").json()
        assert status["state"] == "done"

    def test_empty_data_rejected(self):
        resp = client.post("/jobs", json={
            "data": [],
            "taxonomy": {},
        })
        assert resp.status_code == 400

    def test_file_download_with_valid_signature(self):
        """Verify signed URL download works."""
        clean_rows = [
            {
                "record_id": "DL-001",
                "date": "2025-01-15",
                "amount": 100.0,
                "quantity": 5,
                "unit_price": 20.0,
                "unit": "EA",
                "supplier_name": "Caterpillar",
                "category_l1": "Equipment & Parts",
                "description": "Test download",
            }
        ]
        resp = client.post("/jobs", json={
            "data": clean_rows,
            "taxonomy": SAMPLE_TAXONOMY,
            "vendor_dictionary": {"Caterpillar": "Caterpillar"},
        })
        job_id = resp.json()["job_id"]

        results = client.get(f"/jobs/{job_id}/results").json()
        for name, url in results["signed_urls"].items():
            dl_resp = client.get(url)
            assert dl_resp.status_code == 200, f"Download failed for {name}: {dl_resp.status_code}"

    def test_convergence_trace_recorded(self):
        """Verify convergence trace captures iteration data."""
        dirty = _generate_dirty_rows(30)
        resp = client.post("/jobs", json={
            "data": dirty,
            "taxonomy": SAMPLE_TAXONOMY,
            "abbreviation_dictionary": SAMPLE_ABBREV_DICT,
            "vendor_dictionary": SAMPLE_VENDOR_DICT,
            "max_iterations": 3,
            "llm_batch_size": 50,
        })
        job_id = resp.json()["job_id"]

        for _ in range(5):
            status = client.get(f"/jobs/{job_id}/status").json()
            if status["state"] == "done":
                break
            if status["state"] != "awaiting_llm_batch":
                continue

            batch_resp = client.get(f"/jobs/{job_id}/next_llm_batch")
            if batch_resp.status_code == 204:
                continue

            batch = batch_resp.json()["batch"]
            answers = mock_llm_respond(batch)
            client.post(f"/jobs/{job_id}/llm_batch_response", json={"responses": answers})

        job = job_store.get_job(job_id)
        assert len(job["convergence_trace"]) >= 1
        for entry in job["convergence_trace"]:
            assert "iteration" in entry
            assert "flagged_count" in entry
            assert "timestamp" in entry


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
