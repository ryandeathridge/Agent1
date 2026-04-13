"""Assembles xlsx output files with openpyxl color coding.

Produces four artefacts:
  1. cleaned_data.xlsx       — all rows, color-coded cells
  2. changes_only.xlsx       — rows that changed, with before/after columns
  3. user_verification_required.xlsx — red rows with reason per field
  4. report_summary.json     — counts, timings, rule stats, convergence trace

Color scheme (PatternFill):
  Green  FFB6D7A8 — cell changed by deterministic or LLM rule
  Yellow FFFFF2CC — LLM returned low confidence (< 0.7)
  Red    FFE06666 — could not clean, user review needed

Streams writes row-by-row via openpyxl's write-only optimised mode so the
full dataset never has to fit in memory at once.
"""

from __future__ import annotations

import json
import os
import time
from typing import Any, Dict, List, Optional, Set, Tuple

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Color fills
# ---------------------------------------------------------------------------

FILL_GREEN = PatternFill(start_color="FFB6D7A8", end_color="FFB6D7A8", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")
FILL_RED = PatternFill(start_color="FFE06666", end_color="FFE06666", fill_type="solid")
HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, color="FFFFFF")

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _change_index(changes: List[dict]) -> Dict[int, Dict[str, dict]]:
    """Build {row_id: {field: change_record}} for fast lookup."""
    idx: Dict[int, Dict[str, dict]] = {}
    for c in changes:
        rid = c.get("row_id")
        if rid is None:
            continue
        idx.setdefault(rid, {})
        idx[rid][c["field"]] = c
    return idx


def _flagged_index(flagged_rows: List[dict]) -> Dict[int, dict]:
    """Build {row_id: flagged_record} for fast lookup."""
    return {f["row_id"]: f for f in flagged_rows}


def _write_header(ws, columns: List[str]) -> None:
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _auto_width(ws, columns: List[str]) -> None:
    for col_idx, col_name in enumerate(columns, 1):
        width = max(len(str(col_name)) + 2, 12)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width, 40)


# ---------------------------------------------------------------------------
# 1. cleaned_data.xlsx
# ---------------------------------------------------------------------------


def build_cleaned_data(
    rows: List[Dict[str, Any]],
    changes: List[dict],
    flagged_rows: List[dict],
    output_path: str,
) -> str:
    change_idx = _change_index(changes)
    flag_idx = _flagged_index(flagged_rows)

    if not rows:
        columns = ["(no data)"]
    else:
        columns = list(rows[0].keys())

    wb = Workbook()
    ws = wb.active
    ws.title = "Cleaned Data"
    _write_header(ws, columns)

    for row_i, row in enumerate(rows):
        row_changes = change_idx.get(row_i, {})
        is_flagged = row_i in flag_idx
        flagged_fields = set(flag_idx[row_i].get("fields", {}).keys()) if is_flagged else set()

        for col_idx, col_name in enumerate(columns, 1):
            value = row.get(col_name)
            cell = ws.cell(row=row_i + 2, column=col_idx, value=value)

            if col_name in flagged_fields:
                cell.fill = FILL_RED
            elif col_name in row_changes:
                c = row_changes[col_name]
                conf = c.get("confidence", 1.0)
                if c.get("rule") == "llm_response" and conf < 0.7:
                    cell.fill = FILL_YELLOW
                else:
                    cell.fill = FILL_GREEN

    _auto_width(ws, columns)
    wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# 2. changes_only.xlsx
# ---------------------------------------------------------------------------


def build_changes_only(
    rows: List[Dict[str, Any]],
    original_rows: List[Dict[str, Any]],
    changes: List[dict],
    output_path: str,
) -> str:
    change_idx = _change_index(changes)
    changed_row_ids = sorted(change_idx.keys())

    if not rows:
        columns = ["(no data)"]
    else:
        columns = list(rows[0].keys())

    before_after_cols = []
    for c in columns:
        before_after_cols.append(c)
        before_after_cols.append(f"{c}__before")

    wb = Workbook()
    ws = wb.active
    ws.title = "Changes Only"
    _write_header(ws, ["row_id"] + before_after_cols)

    excel_row = 2
    for rid in changed_row_ids:
        if rid >= len(rows):
            continue
        row = rows[rid]
        orig = original_rows[rid] if rid < len(original_rows) else {}
        row_changes = change_idx.get(rid, {})

        ws.cell(row=excel_row, column=1, value=rid)
        col_offset = 2
        for col_name in columns:
            current_val = row.get(col_name)
            original_val = orig.get(col_name)

            cell_current = ws.cell(row=excel_row, column=col_offset, value=current_val)
            cell_before = ws.cell(row=excel_row, column=col_offset + 1, value=original_val)

            if col_name in row_changes:
                cell_current.fill = FILL_GREEN
            col_offset += 2

        excel_row += 1

    _auto_width(ws, ["row_id"] + before_after_cols)
    wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# 3. user_verification_required.xlsx
# ---------------------------------------------------------------------------


def build_verification_required(
    rows: List[Dict[str, Any]],
    flagged_rows: List[dict],
    output_path: str,
) -> str:
    if not rows:
        columns = ["(no data)"]
    else:
        columns = list(rows[0].keys())

    reason_columns = ["_flag_fields", "_flag_reasons"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Verification Required"
    _write_header(ws, columns + reason_columns)

    excel_row = 2
    for flagged in flagged_rows:
        rid = flagged["row_id"]
        if rid >= len(rows):
            continue
        row = rows[rid]
        reasons = flagged.get("reasons", {})
        flagged_fields = set(flagged.get("fields", {}).keys())

        for col_idx, col_name in enumerate(columns, 1):
            value = row.get(col_name)
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            if col_name in flagged_fields:
                cell.fill = FILL_RED

        ws.cell(row=excel_row, column=len(columns) + 1, value=", ".join(flagged_fields))
        ws.cell(row=excel_row, column=len(columns) + 2, value=json.dumps(reasons))
        excel_row += 1

    _auto_width(ws, columns + reason_columns)
    wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# 4. report_summary.json
# ---------------------------------------------------------------------------


def build_report_summary(job: Dict[str, Any], output_path: str) -> str:
    summary = {
        "job_id": job.get("job_id"),
        "total_rows": job.get("total_rows", 0),
        "rows_processed": job.get("rows_processed", 0),
        "flagged_remaining": job.get("flagged_count", 0),
        "total_changes": len(job.get("changes", [])),
        "iterations": job.get("iteration", 0),
        "rule_stats": job.get("rule_stats", {}),
        "pass_timings": job.get("pass_timings", []),
        "convergence_trace": job.get("convergence_trace", []),
        "llm_batch_count": len([
            t for t in job.get("convergence_trace", [])
        ]),
        "created_at": job.get("created_at"),
        "completed_at": time.time(),
    }
    with open(output_path, "w") as f:
        json.dump(summary, f, indent=2, default=str)
    return output_path


# ---------------------------------------------------------------------------
# Master builder
# ---------------------------------------------------------------------------


def build_all_outputs(job: Dict[str, Any], output_dir: str) -> Dict[str, str]:
    """Build all 4 output files and return {name: path} dict."""
    os.makedirs(output_dir, exist_ok=True)

    rows = job.get("cleaned_data", [])
    original = job.get("original_data", [])
    changes = job.get("changes", [])
    flagged = job.get("flagged_rows", [])

    paths = {}
    paths["cleaned_data"] = build_cleaned_data(
        rows, changes, flagged,
        os.path.join(output_dir, "cleaned_data.xlsx"),
    )
    paths["changes_only"] = build_changes_only(
        rows, original, changes,
        os.path.join(output_dir, "changes_only.xlsx"),
    )
    paths["user_verification_required"] = build_verification_required(
        rows, flagged,
        os.path.join(output_dir, "user_verification_required.xlsx"),
    )
    paths["report_summary"] = build_report_summary(
        job,
        os.path.join(output_dir, "report_summary.json"),
    )

    return paths
